"""Excel export for recommendations — reusable, UI-independent.

Designed as a thin layer over openpyxl so it can be called from:

1. **Streamlit** — `app.py` uses ``recommendations_to_excel_bytes`` for the
   "Download as Excel" button (``st.download_button``).
2. **Batch CLI** — a future ``batch_eval.py`` script can use
   ``append_recommendations_to_excel`` to walk a folder of .docx files
   and write one consolidated workbook.
3. **RPA bots** (UiPath, Power Automate) — the Streamlit "Download as
   Excel" button is fully clickable by browser-automation bots, so the
   same code path serves both human and robot users.

Sheet layout (one sheet, one row per recommendation)
-----------------------------------------------------
| Section | Columns |
|---|---|
| **Case-level** (repeated per recommendation) | case_id · run_date · client_situation · safety_verdict · routing_category · complexity · sub_needs · n_candidates_considered |
| **Recommendation-level** (1 row each) | rec_rank · programme · kind · fit_score · rationale · evidence_quote · eligibility_flags · url · retrieval_score |

This denormalised shape is intentional — it's the easiest format for an
RPA loop ("for each row, do X") and for downstream pivot-table analysis.
"""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Iterable, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Import for type hint only — actual import is deferred to avoid a circular
# dependency when recommender.py imports this module.
try:
    from recommender import RecommendationResponse
except ImportError:  # pragma: no cover - only triggers under unusual layouts
    RecommendationResponse = "RecommendationResponse"  # type: ignore


HEADERS: list[tuple[str, int]] = [
    # (column header, width in chars)
    ("case_id",                   12),
    ("run_date",                  20),
    ("client_situation",          50),
    ("safety_verdict",            14),
    ("routing_category",          18),
    ("complexity",                12),
    ("sub_needs",                 40),
    ("n_candidates_considered",   10),
    ("rec_rank",                   8),
    ("programme",                 32),
    ("kind",                      10),
    ("fit_score",                  9),
    ("rationale",                 60),
    ("evidence_quote",            60),
    ("eligibility_flags",         30),
    ("url",                       40),
    ("retrieval_score",           12),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FIT_FILLS = {
    5: PatternFill("solid", fgColor="C6EFCE"),  # green
    4: PatternFill("solid", fgColor="E2EFDA"),
    3: PatternFill("solid", fgColor="FFF2CC"),  # yellow
    2: PatternFill("solid", fgColor="FCE4D6"),
    1: PatternFill("solid", fgColor="FFC7CE"),  # red
}


def _response_to_rows(
    response: "RecommendationResponse", case_id: str = ""
) -> list[list]:
    """Flatten one ``RecommendationResponse`` into N denormalised rows.

    If ``response.blocked`` is True, returns a single row with the block
    reason in the rationale column.
    """
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    routing = (
        (response.classification or {}).get("category", "")
        if response.classification else ""
    )
    complexity = (
        "complex" if (response.decomposition or {}).get("is_complex")
        else "simple"
    )
    sub_needs = " | ".join(
        (response.decomposition or {}).get("sub_needs", []) or []
    )

    case_cols = [
        case_id,
        run_date,
        response.client_situation,
        "blocked" if response.blocked else "safe",
        routing,
        complexity,
        sub_needs,
        len(response.retrieved_candidates),
    ]

    if response.blocked:
        return [case_cols + [
            0, "", "", 0, response.block_reason, "", "", "", 0.0,
        ]]

    rows: list[list] = []
    for i, rec in enumerate(response.recommendations, start=1):
        rows.append(case_cols + [
            i,
            rec.title,
            rec.kind,
            rec.fit_score,
            rec.rationale,
            rec.evidence_quote,
            " | ".join(rec.eligibility_flags),
            rec.url,
            round(rec.retrieval_score, 4),
        ])
    return rows


def _apply_styles(ws) -> None:
    """Headers, widths, frozen row, fit-score colour bands."""
    fit_score_col = next(
        i for i, (n, _) in enumerate(HEADERS, start=1) if n == "fit_score"
    )

    for col_idx, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True,
            )
        fit_cell = row[fit_score_col - 1]
        try:
            fit = int(fit_cell.value)
            if fit in FIT_FILLS:
                fit_cell.fill = FIT_FILLS[fit]
        except (TypeError, ValueError):
            pass


def recommendations_to_excel_bytes(
    response_or_responses: Union[
        "RecommendationResponse", Iterable["RecommendationResponse"]
    ],
    case_ids: list[str] | None = None,
) -> bytes:
    """Build a fresh in-memory workbook and return its bytes.

    Use this for ``st.download_button`` (Streamlit) — does not touch
    disk.

    Args:
        response_or_responses : One ``RecommendationResponse`` or a list
                                of them.
        case_ids              : Optional list of case IDs matching the
                                responses (one per response). If omitted,
                                IDs are auto-generated as ``C01``, ``C02``…

    Returns:
        ``bytes`` ready to pass to ``st.download_button(data=...)``.
    """
    responses = (
        [response_or_responses]
        if not _is_iterable_of_responses(response_or_responses)
        else list(response_or_responses)
    )
    if case_ids is None:
        case_ids = [f"C{i + 1:02d}" for i in range(len(responses))]

    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"

    ws.append([name for name, _ in HEADERS])

    for resp, cid in zip(responses, case_ids):
        for row in _response_to_rows(resp, case_id=cid):
            ws.append(row)

    _apply_styles(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def append_recommendations_to_excel(
    response: "RecommendationResponse",
    output_path: Union[str, Path],
    case_id: str = "",
) -> Path:
    """Append rows for ``response`` to ``output_path``.

    Creates the file (with headers) if it does not exist. This is the
    function a batch CLI or RPA bot would call once per case.

    Args:
        response    : One ``RecommendationResponse``.
        output_path : Target .xlsx path. Created if missing.
        case_id     : Optional ID for this case (e.g., the source .docx
                      filename, or a sequence like ``Case_01``).

    Returns:
        The resolved Path that was written to.
    """
    path = Path(output_path)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Recommendations"
        ws.append([name for name, _ in HEADERS])

    for row in _response_to_rows(response, case_id=case_id):
        ws.append(row)

    _apply_styles(ws)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_iterable_of_responses(obj) -> bool:
    """Heuristic — distinguish one response from a list/tuple of them."""
    if hasattr(obj, "recommendations") and hasattr(obj, "client_situation"):
        return False  # it's a single RecommendationResponse
    try:
        iter(obj)
        return True
    except TypeError:
        return False


# ---------------------------------------------------------------------------
# Smoke test — `python excel_export.py`
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    from recommender import recommend

    out = Path(__file__).parent / "samples" / "smoke_test.xlsx"
    out.parent.mkdir(exist_ok=True)

    print("Running recommend() for smoke test...")
    response = recommend(
        "Single mother of two young children, recently lost her job, "
        "needs financial help with rent and utilities."
    )
    print(f"  Got {len(response.recommendations)} recommendations.")

    print(f"Writing to {out} ...")
    append_recommendations_to_excel(response, out, case_id="SMOKE_01")
    print(f"  Wrote {out}  ({out.stat().st_size} bytes)")

    print("\nGenerating in-memory bytes (Streamlit path)...")
    blob = recommendations_to_excel_bytes(response, case_ids=["SMOKE_BYTES"])
    print(f"  Got {len(blob):,} bytes of XLSX.")

    print("\nExcel export OK.")
